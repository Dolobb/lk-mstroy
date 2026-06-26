# -*- coding: utf-8 -*-
"""Build the result workbook in the style of шаблон.xlsx from cached NPS TIS data.
Per vehicle / day / shift: engine work time (hh:mm) | fuel consumed (L).
"""
import json, os, datetime as dt
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

CACHE = "_local/comparison-cache"
OUT = "Показатели НПС-ТИС 20.05-20.06.2026.xlsx"

VEHICLES = [
    ("0097081",   37498, "МО-10",  "AIRMAN PDS390SC-W Компрессорная установка"),
    ("5578",      33251, "МО-90",  "BOREY 55-7F Компрессорная установка"),
    ("О415МУ977", 20273, "АТС",    "FAW J6 Автосамосвал"),
    ("20015034",  40203, "МО-6",   "JVM G140QW Дизельная электростанция"),
    ("В204РВ790", 20824, "МО-90",  "XCMG XCT30S1 Автокран"),
    ("1042МХ77",  39612, "МО-114", "SANY SCC1000A5 Кран гусеничный"),
    ("3764МР77",  20259, "МО-4",   "XCMG XE210WD Экскаватор колесный"),
    ("9979МО77",  23816, "МО-114", "SANY SCC550A Кран гусеничный"),
    ("7296НА50",  33224, "МО-90",  "XCMG XE210WD Экскаватор колесный"),
]
START = dt.date(2026, 5, 20)
END = dt.date(2026, 6, 20)

DARK = PatternFill("solid", fgColor="FF1F4E78")
MED = PatternFill("solid", fgColor="FF2E75B6")
WHITE_BOLD = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
PLAIN = Font(name="Calibri", size=10, bold=False)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CEN_NOWRAP = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def days():
    d = START
    while d <= END:
        yield d
        d += dt.timedelta(days=1)


def fmt_engine(sec):
    if not sec or sec <= 0:
        return None
    m = int(round(sec / 60.0))
    return f"{m // 60:02d}:{m % 60:02d}"


def read_shift(reg, day, shift):
    """Return (engine_str, fuel_value) or (None, None)."""
    p = os.path.join(CACHE, reg, f"{day.strftime('%Y-%m-%d')}_{shift}.json")
    if not os.path.exists(p):
        return None, None
    try:
        j = json.load(open(p, encoding="utf-8"))
    except Exception:
        return None, None
    if not isinstance(j, dict) or "engineTime" not in j:
        return None, None
    eng = fmt_engine(j.get("engineTime"))
    fuel = None
    fuels = j.get("fuels") or []
    if fuels and fuels[0].get("rate") is not None:
        fuel = round(float(fuels[0]["rate"]), 1)
    return eng, fuel


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "КИП ДСТ"
    daylist = list(days())
    ncols = 3 + len(daylist) * 4
    last_col = get_column_letter(ncols)

    # Row 1-2 title across the data area
    ws.merge_cells(f"D1:{last_col}2")
    t = ws["D1"]
    t.value = "День, смена (время работы двигателя в смену (чч:мм) | расход топлива (л))"
    t.font = Font(name="Calibri", size=16, bold=True)
    t.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(4, ncols + 1):
        ws.cell(row=2, column=c).border = Border(bottom=thin)

    # Left fixed headers (rows 3-5 merged)
    for col, val, w in [(1, "Госномер", 12), (2, "Подразделение", 14), (3, "Марка / модель", 50)]:
        L = get_column_letter(col)
        ws.merge_cells(f"{L}3:{L}5")
        c = ws.cell(row=3, column=col, value=val)
        c.font = WHITE_BOLD; c.fill = DARK; c.alignment = CEN; c.border = BORDER
        ws.column_dimensions[L].width = w

    # Per-day 3-row header
    col = 4
    for d in daylist:
        c1, c4 = get_column_letter(col), get_column_letter(col + 3)
        ws.merge_cells(f"{c1}3:{c4}3")
        h = ws.cell(row=3, column=col, value=d.strftime("%d.%m"))
        h.font = WHITE_BOLD; h.fill = DARK; h.alignment = CEN; h.border = BORDER
        # shift labels (row4): cols col..col+1 = 1 смена, col+2..col+3 = 2 смена
        ws.merge_cells(f"{get_column_letter(col)}4:{get_column_letter(col+1)}4")
        ws.merge_cells(f"{get_column_letter(col+2)}4:{get_column_letter(col+3)}4")
        s1 = ws.cell(row=4, column=col, value="1 смена")
        s2 = ws.cell(row=4, column=col + 2, value="2 смена")
        for s in (s1, s2):
            s.font = WHITE_BOLD; s.fill = MED; s.alignment = CEN
        # row5 sub-headers
        for off, lbl in [(0, "работа двигателя"), (1, "расход топлива"),
                         (2, "работа двигателя"), (3, "расход топлива")]:
            sc = ws.cell(row=5, column=col + off, value=lbl)
            sc.font = WHITE_BOLD; sc.fill = MED; sc.alignment = CEN
        for off in range(4):
            ws.column_dimensions[get_column_letter(col + off)].width = 14
            for rr in (3, 4, 5):
                ws.cell(row=rr, column=col + off).border = BORDER
        col += 4

    ws.row_dimensions[3].height = 20.1
    ws.row_dimensions[4].height = 20.1
    ws.row_dimensions[5].height = 30.0

    # Data rows
    r = 6
    for reg, idmo, pod, model in VEHICLES:
        a = ws.cell(row=r, column=1, value=reg)
        b = ws.cell(row=r, column=2, value=pod)
        m = ws.cell(row=r, column=3, value=model)
        a.font = PLAIN; a.alignment = CEN_NOWRAP; a.border = BORDER
        b.font = PLAIN; b.alignment = CEN_NOWRAP; b.border = BORDER
        m.font = PLAIN; m.alignment = LEFT; m.border = BORDER
        col = 4
        for d in daylist:
            for off, shift in [(0, "morning"), (2, "evening")]:
                eng, fuel = read_shift(reg, d, shift)
                ce = ws.cell(row=r, column=col + off, value=eng)
                cf = ws.cell(row=r, column=col + off + 1, value=fuel)
                for cc in (ce, cf):
                    cc.font = PLAIN; cc.alignment = CEN_NOWRAP; cc.border = BORDER
            col += 4
        r += 1

    ws.freeze_panes = "D6"

    # Описание sheet
    info = wb.create_sheet("Описание")
    rows = [
        ["Источник", "Сторонний TIS НПС (navi.nps-it.ru/api/v3), команда getMonitoringStats"],
        ["Организация", "Дивизион \"Дороги и Мосты\" (ДиМ)"],
        ["Период", "20.05.2026 – 20.06.2026 включительно (32 дня)"],
        ["Смены", "1 смена 07:30–19:30, 2 смена 19:30–07:30 след. дня"],
        ["Часовой пояс", "Москва, UTC+3 (локальный для региона техники)"],
        ["Работа двигателя", "engineTime из getMonitoringStats, формат чч:мм"],
        ["Расход топлива", "fuels[0].rate (литры) за смену"],
        ["Пустая ячейка", "нет работы двигателя / нет данных за смену"],
        ["Сформировано", dt.datetime.now().strftime("%d.%m.%Y %H:%M")],
    ]
    for i, (k, v) in enumerate(rows, 1):
        kc = info.cell(row=i, column=1, value=k); kc.font = Font(bold=True)
        info.cell(row=i, column=2, value=v)
    info.column_dimensions["A"].width = 20
    info.column_dimensions["B"].width = 80

    wb.save(OUT)
    print("saved", OUT)

    # coverage report
    filled = 0; total = 0
    for reg, *_ in VEHICLES:
        for d in daylist:
            for shift in ("morning", "evening"):
                total += 1
                eng, fuel = read_shift(reg, d, shift)
                if eng or fuel is not None:
                    filled += 1
    print(f"coverage: {filled}/{total} shift-cells with engine/fuel data")


if __name__ == "__main__":
    main()
