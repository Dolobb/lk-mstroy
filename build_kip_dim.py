# -*- coding: utf-8 -*-
"""КИП / Под нагрузкой за 20.05-20.06.2026 для 5 машин ДиМ по новым нормам.
Формат — точно по kip-dim-2026-05-20--2026-06-20.xlsx. Источник — comparison-cache/.
Формулы — kip/server/src/services/kpiCalculator.ts.
"""
import json, os, datetime as dt
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

CACHE = "_local/comparison-cache"
OUT = "КИП-ДиМ-новые-нормы-2026-05-20--2026-06-20.xlsx"
START = dt.date(2026, 5, 20)
END = dt.date(2026, 6, 20)
SHIFT_SEC = 12 * 3600
STAY_HOURS = 12.0

# (regNumber, Подразделение, Марка/модель, норма л/ч) — порядок и тексты со скриншота
VEHICLES = [
    ("В204РВ790", "МО-90",  "XCMG XCT30S1 Автокран",         6.5),
    ("1042МХ77",  "МО-114", "SANY SCC1000A5 Кран гусеничный", 6.5),
    ("3764МР77",  "МО-4",   "XCMG XE210WD Экскаватор колесный", 16.0),
    ("9979МО77",  "МО-114", "SANY SCC550A Кран гусеничный",   7.0),
    ("7296НА50",  "МО-90",  "XCMG XE210WD Экскаватор колесный", 16.0),
]

DARK = PatternFill("solid", fgColor="1F4E78")
GREEN = PatternFill("solid", fgColor="375623")
MED = PatternFill("solid", fgColor="2E75B6")
WB = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
TITLEF = Font(name="Calibri", size=11, bold=True)
PLAIN = Font(name="Calibri", size=10)
CEN = Alignment("center", "center", wrap_text=True)
CENp = Alignment("center", "center")
LEFT = Alignment("left", "center", wrap_text=True)
thin = Side(style="thin")
BORDER = Border(thin, thin, thin, thin)


def days():
    d = START
    while d <= END:
        yield d
        d += dt.timedelta(days=1)


def load_shift(reg, day, shift):
    p = os.path.join(CACHE, reg, f"{day.strftime('%Y-%m-%d')}_{shift}.json")
    if not os.path.exists(p):
        return None
    try:
        return json.load(open(p, encoding="utf-8"))
    except Exception:
        return None


def ignition_off_in_week(reg, day):
    """True, если в любой смене за предыдущие 7 дней двигатель работал не всю смену."""
    found = False
    for k in range(1, 8):
        d = day - dt.timedelta(days=k)
        for sh in ("morning", "evening"):
            j = load_shift(reg, d, sh)
            if j is None:
                continue
            found = True
            if (j.get("engineTime") or 0) < SHIFT_SEC:
                return True
    return True if not found else False  # нет истории → консервативно True


def calc(reg, day, shift, norm):
    """Возвращает (kip_pct, load_pct) или None, если записи нет."""
    j = load_shift(reg, day, shift)
    if j is None or "engineTime" not in j:
        return None
    et = j.get("engineTime") or 0
    engine_h = et / 3600.0
    f = (j.get("fuels") or [{}])
    f = f[0] if f else {}
    rate = f.get("rate")
    consumed = rate if rate is not None else 0.0
    load_override = kip_override = None
    # Условие 1
    if (rate == 0 or rate is None) and et > 0:
        ac = (f.get("valueBegin") or 0) - (f.get("valueEnd") or 0) \
            + (f.get("charges") or 0) - (f.get("discharges") or 0)
        if ac > 10:
            consumed = ac
        elif ignition_off_in_week(reg, day):
            load_override = 50.0
        else:
            kip_override = 0.0
            load_override = 0.0
    fuel_rate_fact = consumed / engine_h if engine_h > 0 else 0.0
    load = (fuel_rate_fact / norm * 100.0) if norm > 0 else 0.0
    load = min(max(0.0, load), 100.0)
    kip = (min(engine_h / STAY_HOURS, 1.0) * 100.0) if STAY_HOURS > 0 else 0.0
    kip = max(0.0, kip)
    if load_override is not None:
        load = load_override
    if kip_override is not None:
        kip = kip_override
    return kip, load


def pct(v):
    return f"{v:.2f}".replace(".", ",") + "%"


def main():
    daylist = list(days())
    ncol = 6 + len(daylist) * 4   # каждая смена = 2 столбца (КИП | Под нагрузкой)
    last = get_column_letter(ncol)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "КИП ДСТ"

    # R1-2 title
    ws.merge_cells(f"G1:{last}2")
    t = ws["G1"]
    t.value = "День, смена (КИП % / Под нагрузкой %)"
    t.font = TITLEF
    t.alignment = Alignment("left", "center")

    # Левые заголовки A-F (merge R3:R5)
    left_hdr = [("A", "Госномер", DARK, 12), ("B", "Подразделение", DARK, 14),
                ("C", "Марка / модель", DARK, 50), ("D", "Норма расхода, л/ч", DARK, 11),
                ("E", "Средний КИП за месяц", GREEN, 12),
                ("F", "Средний КИП за месяц (без 0/0)", GREEN, 14)]
    for col, txt, fill, w in left_hdr:
        ws.merge_cells(f"{col}3:{col}5")
        c = ws[f"{col}3"]
        c.value = txt; c.font = WB; c.fill = fill; c.alignment = CEN; c.border = BORDER
        ws[f"{col}4"].border = BORDER
        ws[f"{col}5"].border = BORDER
        ws.column_dimensions[col].width = w

    # Даты (R3) + смены (R4) + КИП/Под нагрузкой (R5)
    col = 7
    for d in daylist:
        # дата на 4 столбца
        ws.merge_cells(f"{get_column_letter(col)}3:{get_column_letter(col + 3)}3")
        h = ws.cell(3, col, d.strftime("%d.%m"))
        h.font = WB; h.fill = DARK; h.alignment = CEN; h.border = BORDER
        # смены: 2 столбца каждая
        ws.merge_cells(f"{get_column_letter(col)}4:{get_column_letter(col + 1)}4")
        ws.merge_cells(f"{get_column_letter(col + 2)}4:{get_column_letter(col + 3)}4")
        for soff, lbl in ((0, "1 смена"), (2, "2 смена")):
            sc = ws.cell(4, col + soff, lbl)
            sc.font = WB; sc.fill = MED; sc.alignment = CEN; sc.border = BORDER
        # подзаголовки: КИП | Под нагрузкой | КИП | Под нагрузкой
        for off, lbl in ((0, "КИП"), (1, "Под нагрузкой"), (2, "КИП"), (3, "Под нагрузкой")):
            uc = ws.cell(5, col + off, lbl)
            uc.font = WB; uc.fill = MED; uc.alignment = CEN; uc.border = BORDER
            ws.cell(3, col + off).border = BORDER
            ws.column_dimensions[get_column_letter(col + off)].width = 9
        col += 4

    ws.row_dimensions[3].height = 20.1
    ws.row_dimensions[4].height = 20.1
    ws.row_dimensions[5].height = 30.0

    # Данные (с R6)
    r = 6
    for reg, pod, model, norm in VEHICLES:
        ws.cell(r, 1, reg).alignment = CENp
        ws.cell(r, 2, pod).alignment = CENp
        ws.cell(r, 3, model).alignment = LEFT
        ws.cell(r, 4, norm).alignment = CENp
        kips = []
        kips_nz = []
        col = 7
        for d in daylist:
            for off, sh in ((0, "morning"), (2, "evening")):
                res = calc(reg, d, sh, norm)
                ck = ws.cell(r, col + off)       # КИП
                cl = ws.cell(r, col + off + 1)   # Под нагрузкой
                if res is None:
                    ck.value = cl.value = None
                else:
                    kip, load = res
                    ck.value = pct(kip)
                    cl.value = pct(load)
                    kips.append(kip / 100.0)
                    if not (kip == 0 and load == 0):
                        kips_nz.append(kip / 100.0)
                ck.alignment = CENp; cl.alignment = CENp
            col += 4
        avg = sum(kips) / len(kips) if kips else 0
        avg_nz = sum(kips_nz) / len(kips_nz) if kips_nz else 0
        ec = ws.cell(r, 5, round(avg, 6)); ec.number_format = "0.0%"; ec.alignment = CENp
        fc = ws.cell(r, 6, round(avg_nz, 6)); fc.number_format = "0.0%"; fc.alignment = CENp
        r += 1

    # Шрифты/границы для всей области данных
    for rr in range(6, r):
        for cc in range(1, ncol + 1):
            cell = ws.cell(rr, cc)
            cell.font = PLAIN
            cell.border = BORDER

    ws.freeze_panes = "G6"
    wb.save(OUT)
    print("saved", OUT, "rows", r - 6, "cols", ncol)


if __name__ == "__main__":
    main()
