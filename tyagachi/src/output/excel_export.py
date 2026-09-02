from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter




def _fmt_date(value) -> str:
    """Приводим дату/строку к виду DD.MM.YYYY HH:MM — не пересчитываем, просто текст."""
    if value is None:
        return ""
    
    s = str(value).strip()

    # Если уже в нужном формате — возвращаем как есть
    if not s or s == "None":
        return ""
    
    # Пробуем распарсить ISO-формат (2024-06-01T08:00:00 / 2024-06-01 08:00:00)
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            from datetime import datetime
            dt = datetime.strptime(s[:len(fmt) + 2].rstrip("Z"), fmt)
            return dt.strptime("%d.%m.%Y %H:%M")
        except ValueError:
            continue
    return s  # если не распознали — отдаём строку как есть


def _bold_header(ws, row: int, col: int, text: str):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True, name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color="D9E1F2")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    return cell


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_sheet_request(ws, rec: dict):
    """Лист «Заявка» — общие поля первой записи (все записи дублируют шапку заявки)."""
    ws.title = "Заявка"

    fields = [
        ("Номер заявки",         rec.get("request_number")),
        ("Статус заявки",        rec.get("request_status")),
        ("Статус стабильности",  rec.get("stability_status")),
        ("Адрес отправления",    rec.get("route_start_address")),
        ("Адрес назначения",     rec.get("route_end_address")),
        ("Дата начала маршрута", _fmt_date(rec.get("route_start_date"))),
        ("Дата окончания маршрута", _fmt_date(rec.get("route_end_date"))),
        ("Плановое расстояние (км)", rec.get("route_distance")),
        ("Код объекта затрат",   rec.get("object_expend_code")),
        ("Наименование объекта затрат", rec.get("object_expend_name")),
        ("Наименование груза",   rec.get("order_name_cargo")),
    ]

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 50

    for row_idx, (label, value) in enumerate(fields, start=1):
        _bold_header(ws, row_idx, 1, label)
        cell = ws.cell(row=row_idx, column=2, value=value if value is not None else "")
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _build_sheet_pl(ws, records: list[dict]):
    """Лист «Путевые листы» — таблица, одна строка = один ПЛ."""
    ws.title = "Путевые листы"

    headers = [
        "ПЛ №",
        "Госномер ТС",
        "Название ТС",
        "ID ТС (МО)",
        "Дата выезда (факт)",
        "Дата выезда (план)",
        "Дата возврата (план)",
        "Статус ПЛ",
        "Факт. пробег (км)",
    ]

    col_widths = [14, 16, 30, 14, 22, 22, 22, 20, 22]

    for col_idx, header in enumerate(headers, start=1):
        _bold_header(ws, 1, col_idx, header)
    
    _set_col_widths(ws, col_widths)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for row_idx, r in enumerate(records, start=2):
        row_data = [
            r.get("pl_id"),
            r.get("ts_reg_number"),
            r.get("ts_name_mo"),
            r.get("ts_id_mo"),
            _fmt_date(r.get("pl_date_out")),
            _fmt_date(r.get("pl_date_out_plan")),
            _fmt_date(r.get("pl_date_in_plan")),
            r.get("pl_status"),
            r.get("mon_distance"),
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value is not None else "")
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")


def build_request_workbook(request_number: int, matched_records: list[dict]) -> Workbook:
    """
    Собирает Workbook с двумя листами:
      - «Заявка»         — общие поля заявки
      - «Путевые листы»  — таблица ПЛ
    """
        
    wb = Workbook()

    # Берём первую запись как источник общих полей заявки
    first = matched_records[0] if matched_records else {}

    # Переименовываем/используем дефолтный лист
    ws_req = wb.active
    _build_sheet_request(ws_req, first)

    ws_pl = wb.create_sheet()
    _build_sheet_pl(ws_pl, matched_records)

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
     """Сохраняем Workbook в байты (для StreamingResponse)."""
     buf = BytesIO()
     wb.save(buf)
     buf.seek(0)
     return buf.read()